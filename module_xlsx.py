from openpyxl import Workbook, load_workbook

# Zápis do XLSX
wb = Workbook()
ws = wb.active
ws['A1'] = 'Hello'
ws['B1'] = 'World'
wb.save('example.xlsx')

# Čtení z XLSX
wb = load_workbook('example.xlsx')
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
